from __future__ import annotations

"""
Cache warmer CLI for essential public-finance datasets used by the app.

Writes normalized snapshots under data/cache/ so the app can serve without
relying on live upstreams for every request.

Usage examples:

  python -m services.api.cache_warm plf \
    --base https://data.economie.gouv.fr \
    --dataset plf25-depenses-2025-selon-destination \
    --year 2025

  python -m services.api.cache_warm eurostat-cofog --year 2026 --countries FR,DE,IT
"""

import argparse
import datetime as dt
import csv
import json
import logging
import os
import tempfile
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List

import httpx
from openpyxl import load_workbook

from .clients import eurostat as eu
from .clients import ods


ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DATA_DIR = os.path.join(ROOT, "data")
CACHE_DIR = os.path.join(DATA_DIR, "cache")
LOG = logging.getLogger("cbl.warmers")
DEFAULT_PLF_2026_URL = "https://www.budget.gouv.fr/files/uploads/extract/2024/plf2026/plafonds_missions.xlsx"


def _ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _ods_results(js: Dict[str, Any]) -> List[Dict[str, Any]]:
    # Opendatasoft Explore v2.1 returns a JSON with a top-level `results` list.
    # Be defensive and accept alternative shapes.
    return (
        js.get("results")
        or js.get("records")
        or js.get("data")
        or []
    )


def _slug(s: str) -> str:
    import re
    import unicodedata

    s2 = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s2 = s2.lower()
    s2 = re.sub(r"[^a-z0-9]+", " ", s2)
    return s2


def _guess_fields(meta: Dict[str, Any]) -> Dict[str, str]:
    fields = meta.get("dataset", {}).get("fields") or meta.get("fields") or []
    names = {str(f.get("name") or ""): f for f in fields}
    # Build candidates by slug of name/label
    def cand_score(label: str, name: str, target: str) -> int:
        lx = _slug(label)
        nx = _slug(name)
        score = 0
        if target == "cp":
            for tok in ["cp", "credit", "paiement", "paiements"]:
                if tok in lx or tok in nx:
                    score += 1
        elif target == "ae":
            for tok in ["ae", "autorisation", "engagement"]:
                if tok in lx or tok in nx:
                    score += 1
        elif target == "year":
            for tok in ["exercice", "annee", "year"]:
                if tok in lx or tok in nx:
                    score += 1
        elif target == "mission_code":
            for tok in ["code mission", "mission code", "code_mission"]:
                if tok in lx or tok in nx:
                    score += 1
        elif target == "mission_label":
            for tok in ["mission"]:
                if tok in lx or tok in nx:
                    score += 1
            # Prefer descriptive label columns over type/classification helpers
            if name.lower() in ("mission", "libelle_mission"):
                score += 2
            if name.lower().startswith("type_") or "type" in nx:
                score -= 1
        return score

    def pick(target: str, numeric: bool | None = None) -> str | None:
        best = (0, None)
        for f in fields:
            name = str(f.get("name") or "")
            label = str(f.get("label") or name)
            typ = str(f.get("type") or "")
            if numeric is True and typ not in ("double", "int", "bigint", "float", "decimal"):
                continue
            sc = cand_score(label, name, target)
            if sc > best[0]:
                best = (sc, name)
        return best[1]

    # Prefer explicit known French column names when present
    cp_name = pick("cp", numeric=True) or ("credit_de_paiement" if "credit_de_paiement" in names else None) or "cp"
    ae_name = pick("ae", numeric=True) or ("autorisation_engagement" if "autorisation_engagement" in names else None) or "ae"

    # Mission code/label heuristics
    mission_code = pick("mission_code", numeric=None)
    mission_label = pick("mission_label", numeric=None)
    # Prefer explicit French columns when available
    if not mission_code:
        if "code_mission" in names:
            mission_code = "code_mission"
        elif "mission" in names:
            mission_code = "mission"
    if not mission_label:
        if "libelle_mission" in names:
            mission_label = "libelle_mission"
        elif "mission" in names:
            mission_label = "mission"
    # If label would equal code and a libelle exists, prefer the libelle for label
    if mission_label == mission_code and "libelle_mission" in names:
        mission_label = "libelle_mission"

    return {
        "cp": cp_name,
        "ae": ae_name,
        "year": pick("year", numeric=None) or ("exercice" if "exercice" in names else None),
        "mission_code": mission_code or "mission",
        "mission_label": mission_label or "mission",
    }


def warm_plf_state_budget(
    base: str,
    dataset: str,
    year: int,
    cp_field: str = "",
    ae_field: str = "",
    extra_where: str | None = None,
) -> str:
    """Fetch aggregated PLF/LFI credits by mission and write CSV snapshot.

    Output schema: year, mission_code, mission_label, programme_code, programme_label, cp_eur, ae_eur
    (programme columns left blank at this aggregation level)
    """
    _ensure_dir(CACHE_DIR)
    t0 = time.time()
    LOG.info("[PLF] base=%s dataset=%s year=%s", base, dataset, year)

    # Introspect fields and decide actual names
    meta = ods.dataset_info(base, dataset)
    guesses = _guess_fields(meta)
    cp_col = cp_field or guesses["cp"]
    ae_col = ae_field or guesses["ae"]
    code_col = guesses["mission_code"]
    label_col = guesses["mission_label"]
    year_col = guesses.get("year")

    # Build server-side aggregation; avoid duplicate columns if label==code
    if code_col == label_col:
        select = f"{code_col},sum({cp_col}) as cp_eur,sum({ae_col}) as ae_eur"
        group_by = f"{code_col}"
    else:
        select = f"{code_col},{label_col},sum({cp_col}) as cp_eur,sum({ae_col}) as ae_eur"
        group_by = f"{code_col},{label_col}"
    out_csv = os.path.join(CACHE_DIR, f"state_budget_mission_{year}.csv")

    where = None
    if year_col:
        where = f"{year_col}={year}"
    if extra_where:
        where = f"{where} AND ({extra_where})" if where else extra_where

    rows: List[Dict[str, Any]] = []
    try:
        # Try server-side aggregation first
        js = ods.records(base, dataset, select=select, where=where, group_by=group_by, order_by=code_col, limit=500)
        rows = _ods_results(js)
    except Exception:
        rows = []

    # Fallback: client-side aggregation over rows
    if not rows:
        agg: Dict[str, Dict[str, Any]] = {}
        # Unique selection columns
        base_cols = [code_col, label_col, cp_col, ae_col, year_col or ""]
        uniq_cols: List[str] = []
        for c in base_cols:
            if c and c not in uniq_cols:
                uniq_cols.append(c)
        sel_cols = ",".join(uniq_cols)

        def _parse_conditions(expr: str | None) -> List[tuple[str, str]]:
            if not expr:
                return []
            import re

            conds: List[tuple[str, str]] = []
            # Split on AND (case-insensitive)
            parts = re.split(r"\s+AND\s+", expr, flags=re.IGNORECASE)
            for p in parts:
                m = re.search(r"([A-Za-z0-9_]+)\s*=\s*'([^']*)'", p)
                if not m:
                    m = re.search(r'([A-Za-z0-9_]+)\s*=\s*"([^"]*)"', p)
                if m:
                    conds.append((m.group(1), m.group(2)))
            return conds

        conds = _parse_conditions(extra_where)

        # First try with server-side where; if that errors, fetch without where and filter locally
        tried_without_where = False
        drop_order_by = False
        for attempt in range(3):
            try:
                where_clause = None if tried_without_where else where
                for rec in ods.iterate_records(
                    base,
                    dataset,
                    select=sel_cols,
                    where=where_clause,
                    order_by=None if drop_order_by else code_col,
                    page_size=1000,
                    max_pages=200,
                ):
                    # Local filters
                    if year_col:
                        try:
                            yv = rec.get(year_col)
                            if yv is None:
                                continue
                            # Support numeric/double values
                            if int(float(yv)) != int(year):
                                continue
                        except Exception:
                            continue
                    # Apply simple equality conditions
                    ok = True
                    if conds:
                        for k, v in conds:
                            rv = rec.get(k)
                            if rv is None or str(rv) != v:
                                ok = False
                                break
                    if not ok:
                        continue
                    code = str(rec.get(code_col) or "")
                    label = str(rec.get(label_col) or rec.get(code_col) or "")
                    cpv = float(rec.get(cp_col) or 0)
                    aev = float(rec.get(ae_col) or 0)
                    ent = agg.setdefault(code, {"code": code, "label": label, "cp_eur": 0.0, "ae_eur": 0.0})
                    ent["cp_eur"] = float(ent["cp_eur"]) + cpv
                    ent["ae_eur"] = float(ent["ae_eur"]) + aev
                # If we got here without exception, break
                break
            except Exception:
                # Retry without server-side where
                if not tried_without_where:
                    tried_without_where = True
                elif not drop_order_by:
                    drop_order_by = True
                else:
                    # Already dropped both filters; give up loop
                    break
                continue
        rows = list(agg.values())

        # If nothing matched and we had extra conditions, retry ignoring them (keep year filter only)
        if not rows and conds:
            agg = {}
            try:
                for rec in ods.iterate_records(
                    base,
                    dataset,
                    select=sel_cols,
                    where=None if tried_without_where else where,
                    order_by=None,
                    page_size=1000,
                    max_pages=200,
                ):
                    if year_col:
                        try:
                            yv = rec.get(year_col)
                            if yv is None:
                                continue
                            if int(float(yv)) != int(year):
                                continue
                        except Exception:
                            continue
                    code = str(rec.get(code_col) or "")
                    label = str(rec.get(label_col) or rec.get(code_col) or "")
                    cpv = float(rec.get(cp_col) or 0)
                    aev = float(rec.get(ae_col) or 0)
                    ent = agg.setdefault(code, {"code": code, "label": label, "cp_eur": 0.0, "ae_eur": 0.0})
                    ent["cp_eur"] = float(ent["cp_eur"]) + cpv
                    ent["ae_eur"] = float(ent["ae_eur"]) + aev
            except Exception:
                pass
            rows = list(agg.values())

        # Final safety: if API keeps rejecting even without filters, fetch raw rows (no select/order) and aggregate locally
        if not rows:
            agg = {}
            try:
                for rec in ods.iterate_records(
                    base,
                    dataset,
                    select=None,
                    where=None,
                    order_by=None,
                    page_size=1000,
                    max_pages=200,
                ):
                    # Basic guards: skip rows missing required fields
                    if year_col:
                        try:
                            yv = rec.get(year_col)
                            if yv is None or int(float(yv)) != int(year):
                                continue
                        except Exception:
                            continue
                    code = str(rec.get(code_col) or rec.get("code_mission") or rec.get("mission") or "")
                    label = str(rec.get(label_col) or rec.get("libelle_mission") or rec.get("mission") or code)
                    try:
                        cpv = float(rec.get(cp_col) or rec.get("credit_de_paiement") or 0)
                    except Exception:
                        cpv = 0.0
                    try:
                        aev = float(rec.get(ae_col) or rec.get("autorisation_engagement") or 0)
                    except Exception:
                        aev = 0.0
                    if not code:
                        continue
                    ent = agg.setdefault(code, {"code": code, "label": label, "cp_eur": 0.0, "ae_eur": 0.0})
                    ent["cp_eur"] = float(ent["cp_eur"]) + cpv
                    ent["ae_eur"] = float(ent["ae_eur"]) + aev
            except Exception:
                pass
            rows = list(agg.values())

    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["year", "mission_code", "mission_label", "programme_code", "programme_label", "cp_eur", "ae_eur"])
        for rec in rows:
            code = str(rec.get("code") or rec.get(code_col))
            label = str(rec.get("label") or rec.get(label_col) or rec.get(code_col))
            cp = float(rec.get("cp_eur") or rec.get(cp_col) or 0)
            ae = float(rec.get("ae_eur") or rec.get(ae_col) or 0)
            w.writerow([year, code, label, "", "", cp, ae])
    LOG.info("[PLF] wrote %d rows to %s in %.1fs", len(rows), out_csv, time.time() - t0)
    # Sidecar provenance metadata
    sidecar = {
        "extraction_ts": dt.datetime.now(dt.timezone.utc).isoformat(),
        "base": base,
        "dataset": dataset,
        "year": int(year),
        "where": extra_where or (f"{year_col}={year}" if year_col else None),
        "method": "server_or_client_aggregate",
        "row_count": len(rows),
        "cp_field": cp_col,
        "ae_field": ae_col,
        "mission_code_field": code_col,
        "mission_label_field": label_col,
        "produced_columns": [
            "year",
            "mission_code",
            "mission_label",
            "programme_code",
            "programme_label",
            "cp_eur",
            "ae_eur",
        ],
    }
    with open(out_csv.replace('.csv', '.meta.json'), 'w', encoding='utf-8') as f:
        json.dump(sidecar, f, ensure_ascii=False, indent=2)
    return out_csv


def warm_plf_2026_plafonds(source: str | None = None, output_csv: str | None = None) -> str:
    """Download and normalize the PLF 2026 spending ceilings by mission.

    The official data is only available as PDF/XLSX. We prefer XLSX when present
    and extract a minimal CSV with mission_code, mission_label, and ceiling euros.

    Parameters
    ----------
    source:
        Optional override for the XLSX/PDF URL or local path. When omitted we
        use the `PLF_2026_PLAFONDS_URL` environment variable, falling back to a
        hard-coded default. If the download fails, a bundled sample workbook is
        used so tests remain deterministic.
    output_csv:
        Optional absolute path for the generated CSV. Defaults to
        `data/cache/plf_2026_plafonds.csv`.
    """

    _ensure_dir(CACHE_DIR)
    url = source or os.getenv("PLF_2026_PLAFONDS_URL") or DEFAULT_PLF_2026_URL
    out_path = output_csv or os.path.join(CACHE_DIR, "plf_2026_plafonds.csv")

    tmp_path: str | None = None
    cleanup = False
    try:
        if url.startswith("http://") or url.startswith("https://"):
            LOG.info("[PLF2026] Downloading spending ceilings from %s", url)
            try:
                with httpx.Client(timeout=60.0) as client:
                    resp = client.get(url)
                    resp.raise_for_status()
                    suffix = Path(url).suffix or ".xlsx"
                    fd, tmp_path = tempfile.mkstemp(suffix=suffix)
                    cleanup = True
                    with os.fdopen(fd, "wb") as fh:
                        fh.write(resp.content)
            except Exception as exc:  # pragma: no cover - network dependent
                LOG.warning("[PLF2026] Failed to download %s: %s", url, exc)
                tmp_path = None
        else:
            tmp_path = url if os.path.exists(url) else None

        if not tmp_path or not os.path.exists(tmp_path):
            sample = os.path.join(DATA_DIR, "reference", "plf_2026_plafonds_sample.xlsx")
            if not os.path.exists(sample):
                raise FileNotFoundError("No PLF 2026 ceilings source available")
            LOG.info("[PLF2026] Using bundled sample workbook at %s", sample)
            tmp_path = sample

        rows: List[dict[str, Any]] = []
        wb = load_workbook(tmp_path, data_only=True, read_only=True)
        sheet = wb.active
        header_row_idx = None
        col_map: dict[str, int] = {}
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [str(v).strip() if v is not None else "" for v in row]
            tokens = [v.lower() for v in values]
            if not any(tokens):
                continue
            if header_row_idx is None:
                if any("mission" in t for t in tokens) and any(
                    token for token in tokens if any(k in token for k in ("plafond", "ceiling", "montant"))
                ):
                    for col_idx, token in enumerate(tokens):
                        if "code" in token and "mission" in token:
                            col_map["code"] = col_idx
                        elif "mission" in token and "code" not in token:
                            col_map.setdefault("label", col_idx)
                        elif any(k in token for k in ("plafond", "ceiling", "montant")):
                            col_map["amount"] = col_idx
                    header_row_idx = idx
            else:
                code_idx = col_map.get("code")
                label_idx = col_map.get("label")
                amount_idx = col_map.get("amount")
                if amount_idx is None or (code_idx is None and label_idx is None):
                    continue
                raw_code = str(row[code_idx]).strip() if code_idx is not None and row[code_idx] is not None else ""
                raw_label = str(row[label_idx]).strip() if label_idx is not None and row[label_idx] is not None else ""
                raw_amount = row[amount_idx] if amount_idx is not None else None
                if not raw_code and not raw_label:
                    continue
                # Skip subtotal rows that are clearly aggregates
                if raw_label.lower().startswith("total"):
                    continue

                code = raw_code or ""
                label = raw_label or raw_code
                amount_eur = _parse_plafond_amount(raw_amount)
                if amount_eur is None:
                    continue
                rows.append(
                    {
                        "year": 2026,
                        "mission_code": code,
                        "mission_label": label,
                        "plf_ceiling_eur": amount_eur,
                        "source": url if (url.startswith("http")) else "local",
                    }
                )

        if not rows:
            raise ValueError("No mission rows parsed from PLF 2026 workbook")

        with open(out_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(
                fh,
                fieldnames=["year", "mission_code", "mission_label", "plf_ceiling_eur", "source"],
            )
            writer.writeheader()
            writer.writerows(rows)
    finally:
        if cleanup and tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:  # pragma: no cover - best effort cleanup
                pass

    LOG.info("[PLF2026] Wrote %d mission ceilings to %s", len(rows), out_path)
    return out_path


def _parse_plafond_amount(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        val = float(raw)
    else:
        s = str(raw)
        if not s.strip():
            return None
        s_norm = s.replace("\u202f", "").replace(" ", "").replace(" ", "").replace(",", ".")
        try:
            val = float(s_norm)
        except ValueError:
            return None
    # Official tables report millions of euros; convert to euros
    return val * 1_000_000.0


def warm_eurostat_cofog(year: int, countries: List[str]) -> str:
    """Fetch Eurostat COFOG aggregates and compute shares per country.

    Writes data/cache/eu_cofog_shares_{year}.json with structure:
    { "FR": [{"code":"09","label":"Education","share":0.21}, ...], ... }
    """
    _ensure_dir(CACHE_DIR)
    t0 = time.time()
    LOG.info("[EUROSTAT] shares year=%s in %s", year, ",".join(countries))
    out: Dict[str, Any] = {}
    try:
        js = eu.fetch("gov_10a_exp", {"time": str(year), "unit": "MIO_EUR", "sector": "S13"})
        for c in countries:
            shares = eu.cofog_shares(js, year=year, geo=c)
            if shares:
                out[c] = [{"code": code, "label": label, "share": share} for code, label, share in shares]
    except Exception as e_json:
        out["__warning__"] = (
            "Eurostat fetch failed. Ensure EUROSTAT_BASE is reachable and EUROSTAT_COOKIE is set if required. "
            f"Error: {type(e_json).__name__}"
        )

    # If JSON path yielded nothing for some or all countries, try SDMX-XML per-category fallback
    missing = [c for c in countries if c not in out or not out.get(c)]
    if missing:
        try:
            from .data_loader import _COFOG_LABELS  # type: ignore
            majors = [f"{i:02d}" for i in range(1, 11)]
            for c in missing:
                vals: list[tuple[str, str, float]] = []
                total = 0.0
                for m in majors:
                    key = f"A.MIO_EUR.S13.GF{m}.TE.{c}"
                    v = eu.sdmx_value("gov_10a_exp", key, time=str(year))
                    if v is None:
                        continue
                    total += v
                    label = _COFOG_LABELS.get(m, m)
                    vals.append((m, label, v))
                if total > 0 and vals:
                    vals.sort(key=lambda x: x[2], reverse=True)
                    out[c] = [
                        {"code": code, "label": label, "share": (v / total)} for code, label, v in vals
                    ]
        except Exception:
            pass

    # Final fallback to local mapping
    missing2 = [c for c in countries if c not in out or not out.get(c)]
    if missing2:
        try:
            from .data_loader import allocation_by_cofog
            from .models import Basis

            items = allocation_by_cofog(year, Basis.CP)
            for c in missing2:
                out[c] = [{"code": i.code, "label": i.label, "share": i.share} for i in items]
        except Exception:
            pass
    out_path = os.path.join(CACHE_DIR, f"eu_cofog_shares_{year}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    LOG.info("[EUROSTAT] wrote shares to %s in %.1fs", out_path, time.time() - t0)
    return out_path


# ------------------------------
# Eurostat COFOG subfunction shares (GFxx.y) cache
# ------------------------------

def warm_eurostat_cofog_sub(year: int, countries: List[str]) -> str:
    """Fetch COFOG subfunction values and compute shares relative to total expenditures.

    Writes data/cache/eu_cofog_subshares_{year}.json with structure:
      { "FR": { "07": [{"code":"07.1","label":"...","share":0.025}, ...], ... }, ... }
    """
    _ensure_dir(CACHE_DIR)
    t0 = time.time()
    LOG.info("[EUROSTAT] sub-shares year=%s in %s", year, ",".join(countries))
    out: Dict[str, Any] = {}
    majors = [f"{i:02d}" for i in range(1, 11)]
    try:
        # Fetch per-country with na_item=TE to avoid 404 and reduce payload
        for c in countries:
            js = eu.fetch(
                "gov_10a_exp",
                {"time": str(year), "unit": "MIO_EUR", "sector": "S13", "na_item": "TE", "geo": c},
            )
            dims, _, idx_maps, labels = eu._dim_maps(js)  # type: ignore[attr-defined]
            cof_map = idx_maps.get("cofog99", {})
            # Compute grand total across top-level GFxx for this country
            grand_total = 0.0
            for m in majors:
                v = eu.value_at(js, {"unit": "MIO_EUR", "sector": "S13", "na_item": "TE", "time": str(year), "geo": c, "cofog99": f"GF{m}"})
                if v is not None:
                    grand_total += float(v)
            per_major: Dict[str, List[Dict[str, Any]]] = {}
            for m in majors:
                vals: List[tuple[str, str, float]] = []
                for code in cof_map.keys():
                    if not code.startswith(f"GF{m}") or code == f"GF{m}":
                        continue
                    v = eu.value_at(js, {"unit": "MIO_EUR", "sector": "S13", "na_item": "TE", "time": str(year), "geo": c, "cofog99": code})
                    if v is None:
                        continue
                    lab = labels.get("cofog99", {}).get(code, code)
                    vals.append((code, lab, float(v)))
                if vals and grand_total > 0:
                    arr = []
                    for code, lab, v in sorted(vals, key=lambda x: x[2], reverse=True):
                        share = v / grand_total
                        canon = f"{m}.{code.replace('GF','')[2:]}" if len(code) >= 5 else m
                        arr.append({"code": canon, "label": lab, "share": share})
                    per_major[m] = arr
            if per_major:
                out[c] = per_major
    except Exception as e_json:
        out["__warning__"] = (
            "Eurostat JSON fetch failed for subfunctions; attempting SDMX fallback. "
            f"Error: {type(e_json).__name__}"
        )

    # SDMX fallback for any missing country
    missing = [c for c in countries if c not in out]
    if missing:
        try:
            majors = [f"{i:02d}" for i in range(1, 11)]
            for c in missing:
                # Grand total from top-level majors
                grand_total = 0.0
                for m in majors:
                    v = eu.sdmx_value("gov_10a_exp", f"A.MIO_EUR.S13.GF{m}.TE.{c}", time=str(year))
                    if v is not None:
                        grand_total += float(v)
                per_major: Dict[str, List[Dict[str, Any]]] = {}
                # Known COFOG L2 counts per major (COFOG99)
                cofog_l2_counts: Dict[str, int] = {"01": 7, "02": 4, "03": 7, "04": 9, "05": 6, "06": 6, "07": 7, "08": 4, "09": 6, "10": 9}
                for m in majors:
                    vals: List[tuple[str, float]] = []
                    consecutive_misses = 0
                    max_sub = cofog_l2_counts.get(m, 9)
                    for sub in range(1, max_sub + 1):  # bounded to plausible subcodes
                        code = f"GF{m}{sub}"
                        v = eu.sdmx_value("gov_10a_exp", f"A.MIO_EUR.S13.{code}.TE.{c}", time=str(year))
                        if v is None:
                            consecutive_misses += 1
                            if consecutive_misses >= 2:
                                break
                            continue
                        consecutive_misses = 0
                        vals.append((code, float(v)))
                    if vals and grand_total > 0:
                        arr = []
                        for code, v in sorted(vals, key=lambda x: x[1], reverse=True):
                            share = v / grand_total
                            canon = f"{m}.{code.replace('GF','')[2:]}"
                            arr.append({"code": canon, "label": canon, "share": share})
                        per_major[m] = arr
                if per_major:
                    out[c] = per_major
        except Exception:
            pass

    out_path = os.path.join(CACHE_DIR, f"eu_cofog_subshares_{year}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    LOG.info("[EUROSTAT] wrote sub-shares to %s in %.1fs", out_path, time.time() - t0)
    return out_path
# ------------------------------
# LEGO baseline (expenditures v0)
# ------------------------------

def _load_lego_config() -> Dict[str, Any]:
    path = os.path.join(DATA_DIR, "lego_pieces.json")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _cofog_to_gf(code: str) -> List[str]:
    """Map a COFOG code like '09.1' to Eurostat 'GF091' (and sensible fallbacks).

    Returns a list of candidate codes to try in order.
    """
    code = str(code).strip()
    if not code:
        return []
    cand: List[str] = []
    base = code.split(".")[0]
    sub = code.split(".")[1] if "." in code else None
    if sub and sub != "0":
        cand.append(f"GF{base}{sub}")  # e.g., 09.1 -> GF091
        cand.append(f"GF{base}{sub.zfill(2)}")  # 09.10 -> GF0910 (just in case)
    # top-level
    cand.append(f"GF{base}")
    # Raw as-is (if Eurostat already uses with dot)
    cand.append(code)
    return cand


def _na_item_code(code: str) -> str:
    """Normalize NA_ITEM codes to SDMX (remove dots/hyphens, uppercase)."""
    return str(code).replace(".", "").replace("-", "").upper()


def _na_item_parents(code: str) -> List[str]:
    """Generate fallback NA_ITEM parent codes (e.g., D211 -> D21 -> D2)."""
    c = _na_item_code(code)
    parents = [c]
    # progressively strip trailing characters until length 2 (e.g., D2)
    while len(c) > 2:
        c = c[:-1]
        # stop at boundary where it ends with a digit boundary (e.g., D21 -> D2)
        parents.append(c)
    # de-duplicate while preserving order
    seen: set[str] = set()
    out: List[str] = []
    for x in parents:
        if x not in seen:
            out.append(x)
            seen.add(x)
    return out


def _val_mio(js: Dict[str, Any], year: int, country: str, sector: str, unit: str, cofog_code: str, na_item: str) -> float:
    """Best-effort extraction of a MIO_EUR value for given coordinates.
    Tries several COFOG code candidates.
    """
    # Build base coords; allow missing dims gracefully by value_at
    coords: Dict[str, str] = {"time": str(year)}
    # Always try to set commonly present dims if they exist
    dims = js.get("dimension", {}).get("id") or []
    if "unit" in dims:
        coords["unit"] = unit
    if "geo" in dims:
        coords["geo"] = country
    if "sector" in dims:
        coords["sector"] = sector
    if "na_item" in dims:
        coords["na_item"] = na_item
    # Try COFOG candidates
    from .clients import eurostat as eu_client

    for c in _cofog_to_gf(cofog_code):
        c2 = c
        if "cofog99" in dims:
            coords["cofog99"] = c2
        v = eu_client.value_at(js, coords)
        if v is not None:
            return float(v)
    return 0.0


def warm_lego_baseline(year: int, country: str = "FR", scope: str = "S13") -> str:
    """Compute a baseline by LEGO piece (expenditures v0) and write JSON snapshot.

    Output: data/cache/lego_baseline_{year}.json with fields:
      { year, scope, country, pib_eur, depenses_total_eur, pieces: [{id,type,amount_eur,share}], meta }
    """
    _ensure_dir(CACHE_DIR)
    t0 = time.time()
    LOG.info("[LEGO] build baseline year=%s", year)
    cfg = _load_lego_config()

    # Prepare warning aggregator
    warn_parts: List[str] = []

    # Prefer SDMX-XML for expenditures (more reliable)
    js_exp = None  # legacy JSON disabled in favor of XML
    warn = ""
    # Revenues: use SDMX-XML. We keep a JSON fetch attempt only for diagnostics.
    try:
        js_rev = eu.fetch("gov_10a_main", {"time": str(year), "unit": "MIO_EUR", "geo": country})
    except Exception as e:
        js_rev = {}
        warn_parts.append(f"gov_10a_main JSON failed: {type(e).__name__}")

    # GDP series (for info/ratios)
    try:
        from .data_loader import _read_gdp_series  # type: ignore

        gdp_map = _read_gdp_series()
        pib_eur = float(gdp_map.get(int(year), 0.0))
    except Exception:
        pib_eur = 0.0

    pieces_out: List[Dict[str, Any]] = []
    dep_total = 0.0

    recettes_total = 0.0

    # Detect whether Eurostat expenditure payload looks usable (SDMX shape)
    exp_sdmx = bool(js_exp and isinstance(js_exp.get("dimension"), dict) and js_exp.get("dimension", {}).get("id"))

    # Optional fallback: if no SDMX, derive major COFOG amounts from local mapping (sample/PLF) for the requested year
    major_amounts: Dict[str, float] = {}
    if not exp_sdmx:
        try:
            from .data_loader import allocation_by_cofog  # type: ignore
            from .models import Basis  # type: ignore

            items = allocation_by_cofog(year, Basis("CP"))
            # items have codes like '09' with amount_eur
            major_amounts = {str(i.code): float(i.amount_eur) for i in items}
        except Exception:
            major_amounts = {}

    # ------------------------------
    # Expenditures via bucket allocation (major COFOG x NA_ITEM)
    # ------------------------------
    pieces_cfg = cfg.get("pieces", [])
    # Collect buckets that appear in config
    majors: set[str] = set()
    na_set: set[str] = set()
    for p in pieces_cfg:
        if str(p.get("type")) != "expenditure":
            continue
        for mc in (p.get("mapping", {}).get("cofog") or []):
            m = str(mc.get("code", "")).split(".")[0]
            if m:
                majors.add(m)
        for ni in (p.get("mapping", {}).get("na_item") or []):
            na_set.add(_na_item_code(str(ni.get("code", ""))))
    # Fetch bucket totals once (MIO_EUR)
    bucket_totals: Dict[tuple[str, str], float] = {}
    for m in majors:
        gf = f"GF{m}"
        for na in na_set:
            key = f"A.MIO_EUR.S13.{gf}.{na}.{country}"
            val = eu.sdmx_value("gov_10a_exp", key, time=str(year)) or 0.0
            bucket_totals[(m, na)] = float(val)

    # Compute piece weights per bucket and allocate
    exp_amounts: Dict[str, float] = {str(p.get("id")): 0.0 for p in pieces_cfg if str(p.get("type")) == "expenditure"}
    for (m, na), total_mio in bucket_totals.items():
        if total_mio <= 0.0:
            continue
        # Collect weights across pieces for this bucket
        weights: Dict[str, float] = {}
        w_sum = 0.0
        for p in pieces_cfg:
            if str(p.get("type")) != "expenditure":
                continue
            pid = str(p.get("id"))
            cofogs = (p.get("mapping", {}).get("cofog") or [])
            nas = (p.get("mapping", {}).get("na_item") or [])
            w_cof = 0.0
            for mc in cofogs:
                if str(mc.get("code", "")).split(".")[0] == m:
                    w_cof += float(mc.get("weight", 1.0))
            w_na = 0.0
            for ni in nas:
                if _na_item_code(str(ni.get("code", ""))) == na:
                    w_na += float(ni.get("weight", 1.0))
            w = w_cof * w_na
            if w > 0.0:
                weights[pid] = w
                w_sum += w
        if w_sum <= 0.0:
            continue
        # Allocate MIO_EUR total to pieces by normalized weights
        for pid, w in weights.items():
            share = w / w_sum
            exp_amounts[pid] = exp_amounts.get(pid, 0.0) + (total_mio * share * 1_000_000.0)

    # Fill debt_interest from COFOG 01.7 total (TE), since D.41 is not exposed in gov_10a_exp
    try:
        key_di = f"A.MIO_EUR.{scope}.GF0107.TE.{country}"
        di_mio = eu.sdmx_value("gov_10a_exp", key_di, time=str(year))
        if di_mio is None:
            di_mio = eu.sdmx_value("gov_10a_exp", key_di, time=None)
        di_mio = float(di_mio or 0.0)
        if di_mio > 0:
            exp_amounts["debt_interest"] = di_mio * 1_000_000.0
            warn_parts.append("debt_interest from COFOG 01.7 TE (D.41 not exposed in gov_10a_exp)")
    except Exception:
        pass

    # If all zeros, fallback to major-only approximation
    dep_total = sum(exp_amounts.values())
    if dep_total <= 0.0 and major_amounts:
        for p in pieces_cfg:
            if str(p.get("type")) != "expenditure":
                continue
            pid = str(p.get("id"))
            approx = 0.0
            for mc in (p.get("mapping", {}).get("cofog") or []):
                major = str(mc.get("code", "")).split(".")[0]
                w = float(mc.get("weight", 1.0))
                approx += w * float(major_amounts.get(major, 0.0))
            exp_amounts[pid] = approx
        dep_total = sum(exp_amounts.values())

    # Helper: SDMX XML with fallback to last available if the requested year has no Obs
    def _sdmx_value_fallback(flow: str, key: str, y: int) -> float:
        v = eu.sdmx_value(flow, key, time=str(y))
        if v is None:
            v = eu.sdmx_value(flow, key, time=None)
        return float(v or 0.0)

    # Load configurable revenue splits (with sane defaults)
    def _load_revenue_splits() -> Dict[str, Any]:
        try:
            path = os.path.join(DATA_DIR, "revenue_splits.json")
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        # Defaults match initial documented assumptions
        return {
            "vat": {"standard": 0.70, "reduced": 0.30},
            "income_tax": {"pit": 0.60, "cit": 0.40},
            "d29": {"wage_tax": 0.14, "env": 0.10, "fines": 0.02, "transfers": 0.24},
        }

    splits_cfg = _load_revenue_splits()
    VAT_STANDARD_SPLIT = float(splits_cfg.get("vat", {}).get("standard", 0.70))
    PIT_SPLIT = float(splits_cfg.get("income_tax", {}).get("pit", 0.60))
    D29_WAGE = float(splits_cfg.get("d29", {}).get("wage_tax", 0.14))
    D29_ENV = float(splits_cfg.get("d29", {}).get("env", 0.10))
    D29_FINES = float(splits_cfg.get("d29", {}).get("fines", 0.02))
    D29_TRANSFERS = float(splits_cfg.get("d29", {}).get("transfers", 0.24))
    D29_OTHER = max(0.0, 1.0 - (D29_WAGE + D29_ENV + D29_FINES + D29_TRANSFERS))

    # Pre-fetch main revenue bases in MIO_EUR
    # gov_10a_taxag exposes taxes/social contributions by ESA code
    taxag_codes = [
        "D211",  # VAT
        "D214A", "D214B", "D214C",  # excises
        "D29",  # other taxes on production (for splits)
        "D59A",  # recurrent property taxes
        "D51",   # taxes on income etc. (split PIT/CIT)
        "D611", "D612", "D613",  # social contributions
    ]
    taxag_vals: Dict[str, float] = {}
    for c in taxag_codes:
        taxag_vals[c] = _sdmx_value_fallback("gov_10a_taxag", f"A.MIO_EUR.{scope}.{c}.{country}", year)

    # gov_10a_main exposes sales/service revenue and totals
    main_codes = ["P11", "P12"]
    main_vals: Dict[str, float] = {}
    for c in main_codes:
        main_vals[c] = _sdmx_value_fallback("gov_10a_main", f"A.MIO_EUR.{scope}.{c}.{country}", year)

    # Splits above may come from config; ensure residual share for generic D29 (if used)

    # Build pieces_out with expenditures amounts
    for p in pieces_cfg:
        pid = str(p.get("id"))
        ptype = str(p.get("type"))
        amt_eur: float | None = None
        if ptype == "expenditure":
            amt_eur = float(exp_amounts.get(pid, 0.0))
            dep_total += 0.0  # already summed
        elif ptype == "revenue":
            pid = str(p.get("id"))
            esa = p.get("mapping", {}).get("esa") or []
            total_mio = 0.0
            for ent in esa:
                code_raw = str(ent.get("code"))
                w = float(ent.get("weight", 1.0))
                code = _na_item_code(code_raw)
                base = code
                ratio = 1.0
                flow = "taxag"  # shorthand for gov_10a_taxag
                # Map pseudo-codes and choose base/ratio
                if code in ("P11", "P12"):
                    flow = "main"
                    base = code
                elif code == "D211":
                    base = "D211"
                    # Split by piece id into standard/reduced
                    if pid == "rev_vat_standard":
                        ratio = VAT_STANDARD_SPLIT
                    elif pid == "rev_vat_reduced":
                        ratio = 1.0 - VAT_STANDARD_SPLIT
                elif code.startswith("D51_"):
                    base = "D51"
                    if code.endswith("PIT"):
                        ratio = PIT_SPLIT
                    elif code.endswith("CIT"):
                        ratio = 1.0 - PIT_SPLIT
                elif code.startswith("D29_"):
                    base = "D29"
                    if code.endswith("WAGE_TAX"):
                        ratio = D29_WAGE
                    elif code.endswith("ENV"):
                        ratio = D29_ENV
                    elif code.endswith("FINES"):
                        ratio = D29_FINES
                    elif code.endswith("TRANS"):
                        ratio = D29_TRANSFERS
                elif code == "D29":
                    base = "D29"
                    # Assign only the residual share to the generic D29 piece
                    ratio = D29_OTHER
                elif code == "D59_PROP":
                    base = "D59A"
                elif code == "D59_TRANS":
                    base = "D29"
                    ratio = D29_TRANSFERS
                elif code == "D611_CSG":
                    # CSG/CRDS are not isolated in gov_10a_taxag; skip to avoid double count
                    base = "__NONE__"
                    ratio = 0.0
                # Pull value from the right cache
                if base == "__NONE__":
                    val_mio = 0.0
                else:
                    if flow == "main":
                        val_mio = float(main_vals.get(base, 0.0))
                    else:
                        val_mio = float(taxag_vals.get(base, 0.0))
                total_mio += w * ratio * val_mio
            amt_eur = total_mio * 1_000_000.0
            recettes_total += amt_eur
        pieces_out.append({
            "id": pid,
            "type": ptype,
            "amount_eur": amt_eur,
            "share": None,  # filled for expenditures after total known
        })

    # If fallback path failed to set amounts, try a last-resort approximation using major COFOG totals
    if dep_total == 0.0 and major_amounts:
        for ent, p in zip(pieces_out, cfg.get("pieces", [])):
            if ent.get("type") != "expenditure":
                continue
            cofogs = (p.get("mapping", {}).get("cofog") or [])
            approx = 0.0
            for mc in cofogs:
                major = str(mc.get("code", "")).split(".")[0]
                w = float(mc.get("weight", 1.0))
                approx += w * float(major_amounts.get(major, 0.0))
            ent["amount_eur"] = approx
            dep_total += approx

    # Fill shares for expenditures
    for ent in pieces_out:
        if ent["type"] == "expenditure" and dep_total > 0:
            ent["share"] = float(ent["amount_eur"] or 0.0) / dep_total

    out: Dict[str, Any] = {
        "year": int(year),
        "scope": scope,
        "country": country,
        "pib_eur": pib_eur,
        "depenses_total_eur": dep_total,
        "recettes_total_eur": recettes_total,
        "pieces": pieces_out,
        "meta": {
            "source": "Eurostat SDMX 2.1 (dissemination): gov_10a_exp (exp) + gov_10a_taxag (taxes/contrib) + gov_10a_main (sales/totals)",
            "warning": ("; ".join([w for w in ([warn] + warn_parts) if w]) if (warn or warn_parts) else ""),
        },
    }

    out_path = os.path.join(CACHE_DIR, f"lego_baseline_{year}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    LOG.info("[LEGO] wrote %s (exp=%.0f, rev=%.0f, pieces=%d) in %.1fs", out_path, dep_total, recettes_total, len(pieces_out), time.time() - t0)
    # Sidecar meta for provenance
    sidecar = {
        "extraction_ts": dt.datetime.now(dt.timezone.utc).isoformat(),
        "year": int(year),
        "country": country,
        "scope": scope,
        "method": "Eurostat SDMX 2.1 (gov_10a_exp/taxag/main) with mapping-based fallback",
        "pieces": len(pieces_out),
        "warning": out.get("meta", {}).get("warning", ""),
    }
    with open(out_path.replace('.json', '.meta.json'), 'w', encoding='utf-8') as f:
        json.dump(sidecar, f, ensure_ascii=False, indent=2)
    return out_path


def _main_dup(argv: Iterable[str] | None = None) -> None:
    # Deprecated/unused duplicate CLI retained temporarily during refactor.
    # Intentionally left blank.
    pass


# ------------------------------
# DECP procurement ingestion
# ------------------------------

def warm_decp_procurement(
    year: int,
    csv_path: str | None = None,
    base: str | None = None,
    dataset: str | None = None,
    *,
    enrich_sirene: bool = False,
    sirene_max: int = 100,
    sirene_qps: int = 5,
) -> str:
    """Ingest consolidated DECP-like data (CSV or ODS), deduplicate and roll up lots→contracts.

    Writes: data/cache/procurement_contracts_{year}.csv and a sidecar meta JSON.

    Input expectations (CSV): columns compatible with sample:
      contract_id,buyer_org_id,supplier_siren,supplier_name,signed_date,amount_eur,cpv_code,procedure_type,lot_count,location_code
    If multiple rows share the same (contract_id, signed_date), amounts are summed and lot_count aggregated.
    """
    _ensure_dir(CACHE_DIR)
    t0 = time.time()
    LOG.info("[DECP] start year=%s csv=%s ods=%s:%s enrich_sirene=%s max=%s qps=%s", year, csv_path or '-', base or '-', dataset or '-', enrich_sirene, sirene_max, sirene_qps)

    # Normalize and group by contract_id + signed_date
    def _year_of(s: str | None) -> int | None:
        if not s:
            return None
        try:
            return int(str(s).split("-", 1)[0])
        except Exception:
            return None

    groups: Dict[tuple, Dict[str, Any]] = {}
    def _process_record(rec: Dict[str, Any]) -> None:
        if not rec:
            return
        y = _year_of(rec.get("signed_date") or rec.get("datePublication"))
        if y != year:
            return
        cid = str(rec.get("contract_id") or rec.get("id") or rec.get("id_marche") or rec.get("id_contract") or "").strip()
        if not cid:
            return
        key = (cid, rec.get("signed_date") or rec.get("datePublication") or "")
        ent = groups.setdefault(
            key,
            {
                "contract_id": cid,
                "buyer_org_id": str(rec.get("buyer_org_id") or rec.get("acheteur_id") or ""),
                "supplier_siren": str(rec.get("supplier_siren") or rec.get("siret") or rec.get("siren") or ""),
                "supplier_name": str(rec.get("supplier_name") or rec.get("fournisseur") or rec.get("raisonSociale") or ""),
                "signed_date": str(rec.get("signed_date") or rec.get("datePublication") or ""),
                "amount_eur": 0.0,
                "cpv_code": str(rec.get("cpv_code") or rec.get("cpv") or ""),
                "procedure_type": str(rec.get("procedure_type") or rec.get("procedure") or ""),
                "lot_count": 0,
                "location_code": str(rec.get("location_code") or rec.get("codeCommune") or rec.get("code_postal") or ""),
                "amount_quality": "OK",
            },
        )
        try:
            amt = float(rec.get("amount_eur") or rec.get("montant") or rec.get("valeur") or 0.0)
        except Exception:
            amt = 0.0
        ent["amount_eur"] = float(ent["amount_eur"]) + amt
        try:
            lc = int(rec.get("lot_count") or rec.get("nombreLots") or 1)
        except Exception:
            lc = 1
        ent["lot_count"] = int(ent["lot_count"]) + lc
        if not amt or amt <= 0:
            ent["amount_quality"] = "MISSING"

    # Input sources (in priority): CSV path → ODS → auto-download from data.gouv → sample CSV
    auto_src: str | None = None
    if csv_path:
        import csv as _csv
        with open(csv_path, newline="", encoding="utf-8") as f:
            for rec in _csv.DictReader(f):
                _process_record(rec)
        auto_src = csv_path
    elif base and dataset:
        try:
            cnt = 0
            for item in ods.iterate_records(base, dataset, select=None, where=None, order_by=None, page_size=1000, max_pages=5000):
                _process_record(item)
                cnt += 1
                if cnt % 5000 == 0:
                    LOG.info("[DECP] fetched %d rows...", cnt)
            LOG.info("[DECP] fetched total %d rows from ODS", cnt)
            auto_src = f"ods:{base}:{dataset}"
        except Exception as e:
            LOG.warning("[DECP] ODS fetch failed: %s", e)
    else:
        # Auto-download latest DECP resource (CSV) from data.gouv.fr
        try:
            from .clients import decp as dg
            import httpx as _httpx
            res = dg.latest_resource()
            url = str(res.get("url") or "")
            fmt = str(res.get("format") or "").lower()
            if url and fmt == "csv":
                tmp_csv = os.path.join(CACHE_DIR, f"decp_download_{year}.csv")
                LOG.info("[DECP] downloading %s → %s", url, tmp_csv)
                with _httpx.stream("GET", url, timeout=30.0) as r:
                    r.raise_for_status()
                    with open(tmp_csv, "wb") as out:
                        for chunk in r.iter_bytes(1 << 20):
                            out.write(chunk)
                import csv as _csv
                with open(tmp_csv, newline="", encoding="utf-8") as f:
                    for rec in _csv.DictReader(f):
                        _process_record(rec)
                auto_src = f"datagouv:{url}"
            else:
                LOG.warning("[DECP] No suitable CSV resource found on data.gouv.fr; falling back to sample")
        except Exception as e:
            LOG.warning("[DECP] Auto-download failed: %s", e)
        if not auto_src:
            csv_path = os.path.join(DATA_DIR, "sample_procurement.csv")
            import csv as _csv
            with open(csv_path, newline="", encoding="utf-8") as f:
                for rec in _csv.DictReader(f):
                    _process_record(rec)
            auto_src = csv_path

    # Optional Sirene enrichment (NAF, size) for top suppliers by amount
    supplier_meta: Dict[str, Dict[str, str]] = {}
    if enrich_sirene:
        try:
            # Aggregate amounts by supplier
            sums: Dict[str, float] = {}
            for (_, _), rec in groups.items():
                sir = str(rec.get("supplier_siren") or "")
                if not sir:
                    continue
                try:
                    amt = float(rec.get("amount_eur") or 0.0)
                except Exception:
                    amt = 0.0
                sums[sir] = sums.get(sir, 0.0) + amt
            top = sorted(sums.items(), key=lambda x: x[1], reverse=True)[: max(0, int(sirene_max))]
            sirens = [s for s, _ in top]
            if sirens:
                from .clients import insee as insee_client  # lazy import
                delay = 1.0 / max(1, int(sirene_qps))
                for idx, s in enumerate(sirens, 1):
                    try:
                        js = insee_client.sirene_by_siren(s)
                        ul = js.get("uniteLegale") or js.get("unite_legale") or {}
                        naf = ul.get("activitePrincipaleUniteLegale") or ul.get("activite_principale") or ""
                        size = ul.get("trancheEffectifsUniteLegale") or ul.get("tranche_effectifs") or ""
                        supplier_meta[s] = {"naf": str(naf or ""), "size": str(size or "")}
                    except Exception:
                        continue
                    if delay > 0:
                        time.sleep(delay)
                    if idx % 20 == 0:
                        LOG.info("[DECP] sirene enriched %d/%d", idx, len(sirens))
        except Exception:
            supplier_meta = {}

    out_csv = os.path.join(CACHE_DIR, f"procurement_contracts_{year}.csv")
    import csv as _csv

    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "year",
            "contract_id",
            "buyer_org_id",
            "supplier_siren",
            "supplier_name",
            "signed_date",
            "amount_eur",
            "cpv_code",
            "procedure_type",
            "lot_count",
            "location_code",
            "amount_quality",
            "supplier_naf",
            "supplier_company_size",
        ])
        for (_, _), rec in groups.items():
            sir = str(rec.get("supplier_siren") or "")
            meta = supplier_meta.get(sir, {})
            w.writerow([
                year,
                rec.get("contract_id"),
                rec.get("buyer_org_id"),
                rec.get("supplier_siren"),
                rec.get("supplier_name"),
                rec.get("signed_date"),
                float(rec.get("amount_eur") or 0.0),
                rec.get("cpv_code"),
                rec.get("procedure_type"),
                int(rec.get("lot_count") or 0),
                rec.get("location_code"),
                rec.get("amount_quality"),
                meta.get("naf", ""),
                meta.get("size", ""),
            ])

    # Sidecar metadata
    sidecar = {
        "extraction_ts": dt.datetime.now(dt.timezone.utc).isoformat(),
        "year": int(year),
        "source": auto_src or (csv_path or f"ods:{base}:{dataset}"),
        "row_count": len(groups),
        "note": "Deduplicated by (contract_id, signed_date); lots rolled up by summing amounts and lot_count",
        "sirene_enriched": bool(enrich_sirene and supplier_meta),
        "sirene_enriched_count": len(supplier_meta),
        "produced_columns": [
            "year",
            "contract_id",
            "buyer_org_id",
            "supplier_siren",
            "supplier_name",
            "signed_date",
            "amount_eur",
            "cpv_code",
            "procedure_type",
            "lot_count",
            "location_code",
            "amount_quality",
            "supplier_naf",
            "supplier_company_size",
        ],
    }
    with open(out_csv.replace(".csv", ".meta.json"), "w", encoding="utf-8") as f:
        json.dump(sidecar, f, ensure_ascii=False, indent=2)

    LOG.info("[DECP] wrote %d contracts to %s in %.1fs", len(groups), out_csv, time.time() - t0)
    return out_csv


def main(argv: Iterable[str] | None = None) -> None:
    # Basic CLI logging setup (honors LOG_LEVEL)
    level = getattr(logging, os.getenv("LOG_LEVEL", "INFO").upper(), logging.INFO)
    logging.basicConfig(level=level, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
    # Quiet noisy httpx request logs (4xx expected on some probes)
    logging.getLogger("httpx").setLevel(logging.WARNING)
    logging.getLogger("httpcore").setLevel(logging.WARNING)
    p = argparse.ArgumentParser(description="Cache warmer for essential budget data")
    sub = p.add_subparsers(dest="cmd", required=True)

    # PLF/LFI mission-level credits (ODS)
    sp_plf = sub.add_parser("plf", help="Cache PLF/LFI mission-level credits from ODS")
    sp_plf.add_argument("--base", default="https://data.economie.gouv.fr", help="ODS base URL")
    sp_plf.add_argument("--dataset", required=True, help="Dataset id, e.g. plf25-depenses-2025-selon-destination")
    sp_plf.add_argument("--year", type=int, required=True, help="Budget year (for output tagging)")
    sp_plf.add_argument("--cp-field", default="", help="Field name for CP amount (override autodetect)")
    sp_plf.add_argument("--ae-field", default="", help="Field name for AE amount (override autodetect)")
    sp_plf.add_argument("--where", dest="extra_where", default=None, help="Extra ODS where clause, e.g. typebudget='PLF'")

    # PLF 2026 mission ceilings (PDF/XLSX scraped)
    sp_plf26 = sub.add_parser("plf-2026-plafonds", help="Download PLF 2026 spending ceilings and normalize")
    sp_plf26.add_argument("--source", default=None, help="Override URL or local path to PLF 2026 workbook")
    sp_plf26.add_argument("--output", default=None, help="Optional output CSV path")

    # Eurostat COFOG shares
    sp_eu = sub.add_parser("eurostat-cofog", help="Cache Eurostat COFOG shares for countries/year")
    sp_eu.add_argument("--year", type=int, required=True)
    sp_eu.add_argument("--countries", required=True, help="Comma-separated country codes, e.g. FR,DE,IT")

    # Eurostat COFOG subfunction shares (GFxx.y)
    sp_eu_sub = sub.add_parser("eurostat-cofog-sub", help="Cache Eurostat COFOG subfunction shares for countries/year")
    sp_eu_sub.add_argument("--year", type=int, required=True)
    sp_eu_sub.add_argument("--countries", required=True, help="Comma-separated country codes, e.g. FR,DE,IT")

    # ODS dataset fields helper
    sp_fields = sub.add_parser("ods-fields", help="List fields for an ODS dataset (to help pick cp/ae/year fields)")
    sp_fields.add_argument("--base", default="https://data.economie.gouv.fr")
    sp_fields.add_argument("--dataset", required=True)

    # LEGO baseline warmer (expenditures v0)
    sp_lego = sub.add_parser("lego", help="Build LEGO baseline for a year (expenditures v0)")
    sp_lego.add_argument("--year", type=int, required=True)
    sp_lego.add_argument("--country", default="FR")
    sp_lego.add_argument("--scope", default="S13")

    # DECP procurement ingestion
    sp_decp = sub.add_parser("decp", help="Ingest DECP procurement and write normalized cache")
    sp_decp.add_argument("--year", type=int, required=True)
    sp_decp.add_argument("--csv", dest="csv_path", default=None, help="Path to input CSV (consolidated)")
    sp_decp.add_argument("--base", default=None, help="ODS base URL (optional)")
    sp_decp.add_argument("--dataset", default=None, help="ODS dataset id (optional)")
    sp_decp.add_argument("--enrich-sirene", action="store_true", help="Enrich top suppliers with SIRENE (NAF, size)")
    sp_decp.add_argument("--sirene-max", type=int, default=100, help="Max suppliers to enrich by amount")
    sp_decp.add_argument("--sirene-qps", type=int, default=5, help="Throttle SIRENE lookups (queries per second)")

    # INSEE macro series warmer
    sp_macro = sub.add_parser("macro-insee", help="Warm selected INSEE BDM macro series from a config JSON")
    sp_macro.add_argument("--config", required=True, help="Path to macro series config JSON")

    args = p.parse_args(list(argv) if argv is not None else None)

    if args.cmd == "plf":
        path = warm_plf_state_budget(args.base, args.dataset, args.year, args.cp_field, args.ae_field, args.extra_where)
        print(f"Wrote {path}")
        return

    if args.cmd == "plf-2026-plafonds":
        path = warm_plf_2026_plafonds(args.source, args.output)
        print(f"Wrote {path}")
        return

    if args.cmd == "eurostat-cofog":
        countries = [c.strip() for c in args.countries.split(",") if c.strip()]
        path = warm_eurostat_cofog(args.year, countries)
        print(f"Wrote {path}")
        return

    if args.cmd == "eurostat-cofog-sub":
        countries = [c.strip() for c in args.countries.split(",") if c.strip()]
        path = warm_eurostat_cofog_sub(args.year, countries)
        print(f"Wrote {path}")
        return

    if args.cmd == "ods-fields":
        meta = ods.dataset_info(args.base, args.dataset)
        fields = meta.get("dataset", {}).get("fields") or meta.get("fields") or []
        for f in fields:
            print(f"{f.get('name')}: {f.get('type')} — {f.get('label')}")
        return

    if args.cmd == "lego":
        path = warm_lego_baseline(args.year, country=args.country, scope=args.scope)
        print(f"Wrote {path}")
        return

    if args.cmd == "decp":
        path = warm_decp_procurement(
            args.year,
            csv_path=args.csv_path,
            base=args.base,
            dataset=args.dataset,
            enrich_sirene=bool(getattr(args, "enrich_sirene", False)),
            sirene_max=int(getattr(args, "sirene_max", 100)),
            sirene_qps=int(getattr(args, "sirene_qps", 5)),
        )
        print(f"Wrote {path}")
        return

    if args.cmd == "macro-insee":
        path = warm_macro_insee(args.config)
        print(f"Wrote {path}")
        return


# ------------------------------
# INSEE macro series warmer (deflators, employment)
# ------------------------------

def warm_macro_insee(config_path: str) -> str:
    """Warm selected INSEE BDM series based on a simple config JSON.

    Config format:
    {
      "country": "FR",
      "items": [
        {"id": "deflator_gdp", "dataset": "CNA-2014-PIB", "series": ["PIB-VALUE"]},
        {"id": "employment_total", "dataset": "EMP", "series": ["EMP-POP"]}
      ]
    }
    """
    _ensure_dir(CACHE_DIR)
    t0 = time.time()
    LOG.info("[INSEE] warm macro from %s", config_path)
    import json as _json
    from .clients import insee as insee_client

    with open(config_path, "r", encoding="utf-8") as f:
        cfg = _json.load(f)
    country = cfg.get("country", "FR")
    items = cfg.get("items") or []
    out: dict = {"country": country, "items": []}
    provenance: list[dict] = []
    for it in items:
        ds = str(it.get("dataset"))
        sids = [str(x) for x in (it.get("series") or [])]
        rid = str(it.get("id") or ds)
        try:
            js = insee_client.bdm_series(ds, sids)
        except Exception:
            js = {"error": True}
        out["items"].append({"id": rid, "dataset": ds, "series": sids, "data": js})
        provenance.append({"id": rid, "dataset": ds, "series": sids})
    out_path = os.path.join(CACHE_DIR, f"macro_series_{country}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        _json.dump(out, f, ensure_ascii=False, indent=2)
    # Sidecar
    sidecar = {
        "extraction_ts": dt.datetime.now(dt.timezone.utc).isoformat(),
        "country": country,
        "items": provenance,
        "config": os.path.abspath(config_path),
    }
    with open(out_path.replace(".json", ".meta.json"), "w", encoding="utf-8") as f:
        _json.dump(sidecar, f, ensure_ascii=False, indent=2)
    LOG.info("[INSEE] wrote %s in %.1fs (items=%d)", out_path, time.time() - t0, len(items))
    return out_path



if __name__ == "__main__":
    main()
